from __future__ import annotations

from collections import Counter
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session

from ..domain import redact_text
from ..mitre import technique_label
from .data import ReportData, load_report_data

NAVY = "102A43"
TEAL = "168C8C"
LIGHT = "EEF4F7"
WHITE = "FFFFFF"
MUTED = "5D6B78"
SEVERITY_FILLS = {
    "critical": "8B1E3F",
    "high": "C44536",
    "medium": "D98E04",
    "low": "2D6A9F",
    "informational": "607D8B",
}


class ExcelReportGenerator:
    SHEETS = (
        "Executive Summary",
        "Reconnaissance",
        "Technologies",
        "Vulnerable Components",
        "Endpoints",
        "Parameters",
        "Findings",
        "OWASP Coverage",
        "Attack Paths",
        "Attack Steps",
        "Post-Exploitation Steps",
        "Manual Review",
        "Plugin Health",
        "Attack Path Findings",
        "Assessment Metrics",
    )

    def generate(self, session: Session, assessment_id: str, output_path: Path) -> Path:
        data = load_report_data(session, assessment_id)
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name in self.SHEETS:
            workbook.create_sheet(name)
        self._executive(workbook["Executive Summary"], data)
        self._reconnaissance(workbook["Reconnaissance"], data)
        self._technologies(workbook["Technologies"], data, vulnerable_only=False)
        self._technologies(workbook["Vulnerable Components"], data, vulnerable_only=True)
        self._endpoints(workbook["Endpoints"], data)
        self._parameters(workbook["Parameters"], data)
        self._findings(workbook["Findings"], data)
        self._coverage(workbook["OWASP Coverage"], data)
        self._paths(workbook["Attack Paths"], data)
        self._steps(workbook["Attack Steps"], data)
        self._post_exploitation(workbook["Post-Exploitation Steps"], data)
        self._manual(workbook["Manual Review"], data)
        self._plugins(workbook["Plugin Health"], data)
        self._path_findings(workbook["Attack Path Findings"], data)
        self._metrics(workbook["Assessment Metrics"], data)
        for sheet in workbook.worksheets:
            self._finish_sheet(sheet)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path

    def _executive(self, sheet: Any, data: ReportData) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells("A1:F2")
        sheet["A1"] = "WRAITH CRAWLER - EXECUTIVE SUMMARY"
        sheet["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
        sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
        sheet["A1"].alignment = Alignment(vertical="center")
        metadata = [
            ("Application", data.application_name),
            ("Target", data.target),
            ("Assessment ID", data.assessment_id),
            ("Status", data.status),
            ("Profile", data.profile),
            ("Started", data.started_at),
            ("Completed", data.completed_at),
        ]
        for row, (label, value) in enumerate(metadata, 4):
            sheet.cell(row, 1, label).font = Font(bold=True, color=MUTED)
            sheet.cell(row, 2, value)
            if label in {"Started", "Completed"}:
                sheet.cell(row, 2).number_format = "yyyy-mm-dd hh:mm"
        counts = Counter(f.priority_level for f in data.findings if not f.false_positive)
        sheet["A13"] = "Priority"
        sheet["B13"] = "Count"
        for index, level in enumerate(("critical", "high", "medium", "low", "informational"), 14):
            sheet.cell(index, 1, level.title())
            sheet.cell(index, 2, counts[level])
        self._style_header(sheet, 13, 1, 2)
        sheet["D13"] = "KPI"
        sheet["E13"] = "Value"
        kpis = [
            ("Findings", len(data.findings)),
            ("Attack paths", len(data.attack_paths)),
            ("Manual review backlog", sum(1 for item in data.manual_reviews if item.status == "open")),
            ("Endpoints", len(data.endpoints)),
            ("Technologies", len(data.technologies)),
        ]
        for index, (label, value) in enumerate(kpis, 14):
            sheet.cell(index, 4, label)
            sheet.cell(index, 5, value)
        self._style_header(sheet, 13, 4, 5)
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 44
        sheet.column_dimensions["C"].width = 3
        sheet.column_dimensions["D"].width = 26
        sheet.column_dimensions["E"].width = 14
        sheet.column_dimensions["F"].width = 3

    def _findings(self, sheet: Any, data: ReportData) -> None:
        headers = [
            "Finding ID", "Type", "Family", "Title", "Asset", "Severity", "Priority", "Priority Score",
            "Confidence", "Validation", "Status", "Manual Review", "CWE", "CVE", "OWASP",
            "MITRE ATT&CK", "CVSS", "EPSS",
            "KEV", "Parameters", "Affected Endpoint Count", "Remediation", "First Seen", "Last Seen",
            "Evidence Summary",
            "How an Attacker Exploits", "Capability Gained", "Next Realistic Step",
            "Chain Opportunities", "Confirmed", "Inferred", "Technical Impact",
            "Business Impact", "Attack-Path Break Point",
        ]
        rows = []
        for f in data.findings:
            narrative = f.attacker_narrative or {}
            rows.append([
                f.id, f.finding_type, f.family, f.title, f.asset, f.severity, f.priority_level, f.priority_score,
                f.confidence, f.validation_status, f.status, f.manual_review, ", ".join(f.cwe), ", ".join(f.cve),
                ", ".join(f.owasp), " | ".join(technique_label(item) for item in f.mitre_attack),
                f.cvss, f.epss, f.kev, ", ".join(f.parameters), len(f.affected_endpoints),
                f.remediation, f.first_seen, f.last_seen,
                " | ".join(
                    record.summary + (" [redacted]" if record.sensitive else "")
                    for record in data.evidence
                    if record.finding_id == f.id
                ),
                narrative.get("exploitation", ""),
                narrative.get("capability_gained", ""),
                narrative.get("next_realistic_step", ""),
                " | ".join(narrative.get("chain_opportunities", [])),
                " | ".join(narrative.get("confirmed", [])),
                " | ".join(narrative.get("inferred", [])),
                narrative.get("technical_impact", ""),
                narrative.get("business_impact", ""),
                narrative.get("remediation_break_point", ""),
            ])
        self._write_table(sheet, headers, rows, "FindingsTable")
        for cell in sheet["H"][1:]:
            cell.number_format = "0.0"
        for col in ("Q", "R"):
            for cell in sheet[col][1:]:
                cell.number_format = "0.00"
        for col in ("W", "X"):
            for cell in sheet[col][1:]:
                cell.number_format = "yyyy-mm-dd hh:mm"
        if rows:
            for level, color in SEVERITY_FILLS.items():
                sheet.conditional_formatting.add(
                    f"G2:G{len(rows) + 1}",
                    FormulaRule(formula=[f'$G2="{level}"'], fill=PatternFill("solid", fgColor=color), font=Font(color=WHITE)),
                )

    def _reconnaissance(self, sheet: Any, data: ReportData) -> None:
        headers = [
            "Asset ID", "URL", "Origin", "Scheme", "Hostname", "Port", "Resolved IPs",
            "CNAME", "HTTP Status", "Redirect Chain", "Title", "Server", "CDN/WAF",
            "TLS Details", "Discovery Sources",
        ]
        rows = [
            [
                asset.id, asset.url, asset.origin, asset.scheme, asset.hostname, asset.port,
                ", ".join(asset.resolved_ips), asset.cname or "", asset.http_status,
                " -> ".join(asset.redirect_chain), asset.title or "", asset.server or "",
                asset.cdn_waf or "", self._flat_mapping(asset.tls_details),
                ", ".join(asset.discovery_sources),
            ]
            for asset in data.assets
        ]
        self._write_table(sheet, headers, rows, "ReconnaissanceTable")

    def _endpoints(self, sheet: Any, data: ReportData) -> None:
        parameter_counts = Counter(parameter.endpoint_id for parameter in data.parameters)
        headers = [
            "Endpoint ID", "URL", "Origin", "Path", "Method", "Status", "Content Type",
            "Discovery Sources", "Parameter Count", "Authentication Required",
            "API Classification", "JavaScript Reference", "Confidence",
        ]
        rows = [
            [
                endpoint.id, endpoint.url, endpoint.origin, endpoint.path, endpoint.method,
                endpoint.status_code, endpoint.content_type or "", ", ".join(endpoint.source_plugins),
                parameter_counts[endpoint.id], endpoint.authentication_required,
                endpoint.api_classification or "", endpoint.javascript_source or "", endpoint.confidence,
            ]
            for endpoint in data.endpoints
        ]
        self._write_table(sheet, headers, rows, "EndpointsTable")

    def _parameters(self, sheet: Any, data: ReportData) -> None:
        endpoints = {endpoint.id: endpoint for endpoint in data.endpoints}
        headers = [
            "Parameter ID", "Endpoint", "Method", "Name", "Normalized Name", "Location",
            "Required", "Source", "Risk Categories", "Risk Score", "Sample Metadata",
        ]
        rows = [
            [
                parameter.id,
                endpoints[parameter.endpoint_id].url if parameter.endpoint_id in endpoints else "",
                parameter.method,
                parameter.name,
                parameter.normalized_name,
                parameter.location,
                parameter.required,
                parameter.source,
                ", ".join(parameter.risk_categories),
                parameter.risk_score,
                self._flat_mapping(parameter.sample_metadata),
            ]
            for parameter in data.parameters
        ]
        self._write_table(sheet, headers, rows, "ParametersTable")

    def _coverage(self, sheet: Any, data: ReportData) -> None:
        headers = [
            "Category", "Name", "Status", "Automated Checks", "Plugins", "Tests Attempted",
            "Tests Completed", "Finding Count", "Limitations", "Manual Review Needs",
        ]
        rows = []
        for entry in data.owasp_coverage:
            rows.append([
                entry["category"], entry["name"], entry.get("status", "reference"),
                ", ".join(entry.get("automated_checks", entry.get("capabilities", []))),
                ", ".join(entry.get("plugins", [])), entry.get("tests_attempted", 0),
                entry.get("tests_completed", 0),
                entry.get("findings", sum(1 for finding in data.findings if entry["category"] in finding.owasp)),
                "; ".join(entry.get("limitations", [entry.get("limitation", "")])),
                " | ".join(entry.get("manual_review_needs", [])),
            ])
        self._write_table(sheet, headers, rows, "OWASPCoverageTable")

    def _technologies(self, sheet: Any, data: ReportData, vulnerable_only: bool) -> None:
        headers = ["Technology ID", "Product", "Version", "Category", "Confidence", "Lifecycle State", "Supported", "EOL Date", "Lifecycle Source", "Lifecycle Evidence", "Vulnerability References", "Vulnerability Data", "Detection Source"]
        technologies = [t for t in data.technologies if not vulnerable_only or t.eol_state or t.vulnerability_references]
        rows = [[t.id, t.product, t.version or "", t.category, t.confidence, t.eol_state or "unknown", t.supported, t.eol_date or "", t.lifecycle_source or "", " | ".join(t.lifecycle_evidence), ", ".join(t.vulnerability_references), " | ".join(self._flat_mapping(item) for item in t.vulnerability_data), t.source_plugin] for t in technologies]
        self._write_table(sheet, headers, rows, "VulnerableComponentsTable" if vulnerable_only else "TechnologiesTable")

    def _plugins(self, sheet: Any, data: ReportData) -> None:
        headers = ["Plugin", "Phase", "Security Question", "State", "Failure Reason", "Skip Reason", "Message", "Started", "Completed", "Duration (ms)", "Targets Tested", "Tests Attempted", "Tests Completed", "Findings", "Tool Path", "Tool Version"]
        rows = [[p.plugin_name, p.phase, p.security_question or "", p.state, p.failure_reason or "", p.skip_reason or "", p.message or "", p.started_at, p.completed_at, p.duration_ms, p.targets_tested, p.tests_attempted, p.tests_completed, p.findings_count, p.tool_path or "", p.tool_version or ""] for p in data.plugin_executions]
        self._write_table(sheet, headers, rows, "PluginExecutionTable")
        for col in ("H", "I"):
            for cell in sheet[col][1:]:
                cell.number_format = "yyyy-mm-dd hh:mm"

    def _manual(self, sheet: Any, data: ReportData) -> None:
        headers = ["Review ID", "Status", "Priority", "Candidate Type", "Endpoint", "Parameter", "Reason", "Evidence Summary", "Source Plugin", "Confidence", "Suggested Safe Steps", "Attack Path Relevance", "Resolution"]
        rows = [[m.id, m.status, m.priority, m.candidate_type, m.endpoint, m.parameter or "", m.reason, m.evidence_summary, m.source_plugin, m.confidence, " | ".join(m.suggested_steps), m.attack_path_relevance or "", m.resolution or ""] for m in data.manual_reviews]
        self._write_table(sheet, headers, rows, "ManualReviewTable")

    def _paths(self, sheet: Any, data: ReportData) -> None:
        headers = ["Attack Path ID", "Title", "Score", "Priority", "Confidence", "Classification", "Status", "Critical Path Labels", "MITRE ATT&CK", "Attack Scenario", "Attacker Gain", "Next Step", "Technical Impact", "Business Impact", "Blast Radius", "Evidence Boundary", "Recommended Break Point"]
        rows = [[p.id, p.title, p.score, p.priority, p.confidence, p.classification, p.status, ", ".join(p.critical_path_labels), " | ".join(technique_label(item) for item in p.mitre_attack), p.attack_scenario, p.attacker_gain, p.next_step, p.technical_impact, p.business_impact, p.blast_radius, p.evidence_boundary, p.recommended_break_point] for p in data.attack_paths]
        self._write_table(sheet, headers, rows, "AttackPathsTable")

    def _steps(self, sheet: Any, data: ReportData) -> None:
        nodes = {node.id: node for node in data.attack_nodes}
        headers = ["Attack Path ID", "Step ID", "Source Type", "Source", "Relationship", "Destination Type", "Destination", "Confidence", "Classification", "Rationale", "Evidence Reference", "MITRE ATT&CK"]
        rows = []
        for edge in data.attack_edges:
            source = nodes.get(edge.source_node_id)
            destination = nodes.get(edge.destination_node_id)
            rows.append([edge.attack_path_id, edge.id, source.node_type if source else "", source.label if source else "", edge.relationship, destination.node_type if destination else "", destination.label if destination else "", edge.confidence, edge.classification, edge.rationale, edge.evidence_reference or "", " | ".join(technique_label(item) for item in edge.mitre_attack)])
        self._write_table(sheet, headers, rows, "AttackStepsTable")

    def _post_exploitation(self, sheet: Any, data: ReportData) -> None:
        headers = [
            "Step ID", "Attack Path ID", "Source Finding ID", "Sequence", "Action",
            "Capability", "Classification", "Confidence", "Rationale", "MITRE ATT&CK", "Technical Impact",
            "Business Impact",
        ]
        rows = [
            [
                step.id, step.attack_path_id, step.source_finding_id or "", step.sequence,
                step.action, step.capability, step.classification, step.confidence,
                step.rationale, " | ".join(technique_label(item) for item in step.mitre_attack),
                step.technical_impact or "", step.business_impact or "",
            ]
            for step in data.post_exploitation_steps
        ]
        self._write_table(sheet, headers, rows, "PostExploitationStepsTable")

    def _path_findings(self, sheet: Any, data: ReportData) -> None:
        findings = {finding.id: finding for finding in data.findings}
        headers = ["Attack Path ID", "Finding ID", "Role", "Finding Type", "Title", "Severity", "Priority", "Validation", "MITRE ATT&CK"]
        rows = []
        for link in data.attack_path_findings:
            finding = findings.get(link.finding_id)
            rows.append([link.attack_path_id, link.finding_id, link.role, finding.finding_type if finding else "", finding.title if finding else "", finding.severity if finding else "", finding.priority_level if finding else "", finding.validation_status if finding else "", " | ".join(technique_label(item) for item in finding.mitre_attack) if finding else ""])
        self._write_table(sheet, headers, rows, "AttackPathFindingsTable")

    def _metrics(self, sheet: Any, data: ReportData) -> None:
        headers = ["Metric", "Value", "Recorded At", "Dimensions"]
        rows = [[metric.metric_name, metric.metric_value, metric.recorded_at, "; ".join(f"{key}={value}" for key, value in metric.dimensions.items())] for metric in data.metrics]
        self._write_table(sheet, headers, rows, "AssessmentMetricsTable")
        for cell in sheet["C"][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm"

    @staticmethod
    def _flat_mapping(value: object) -> str:
        if not isinstance(value, dict):
            return str(value)
        return "; ".join(f"{key}={item}" for key, item in value.items())

    def _write_table(self, sheet: Any, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        if not rows:
            sheet.append(["No records"] + [None] * (len(headers) - 1))
        self._style_header(sheet, 1, 1, len(headers))
        end = sheet.max_row
        table = Table(displayName=table_name, ref=f"A1:{get_column_letter(len(headers))}{end}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)

    @staticmethod
    def _style_header(sheet: Any, row: int, first_col: int, last_col: int) -> None:
        for column in range(first_col, last_col + 1):
            cell = sheet.cell(row, column)
            cell.font = Font(name="Aptos", bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(vertical="center")

    def _finish_sheet(self, sheet: Any) -> None:
        sheet.freeze_panes = "A2" if sheet.title != "Executive Summary" else "A4"
        sheet.auto_filter.ref = sheet.dimensions if sheet.title != "Executive Summary" else None
        sheet.sheet_view.showGridLines = False
        thin = Side(style="thin", color="D7E0E5")
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str):
                    cell.value = redact_text(cell.value)
                if cell.font:
                    font = copy(cell.font)
                    font.name = "Aptos"
                    cell.font = font
                else:
                    cell.font = Font(name="Aptos")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row > 1:
                    cell.border = Border(bottom=thin)
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            values = [str(cell.value) for cell in column_cells if cell.value is not None]
            width = min(55, max(11, max((len(value.split("\n")[0]) for value in values), default=10) + 2))
            if any(keyword in str(column_cells[0].value).lower() for keyword in ("description", "remediation", "reason", "impact", "scenario", "boundary", "steps", "endpoint", "target")):
                width = min(55, max(width, 28))
            sheet.column_dimensions[letter].width = width
        sheet.auto_filter.ref = sheet.dimensions if sheet.title != "Executive Summary" else None
