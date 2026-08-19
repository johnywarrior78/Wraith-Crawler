from __future__ import annotations

import httpx
from openpyxl import load_workbook
from pypdf import PdfReader

from wraith_crawler.api import create_app
from wraith_crawler.reporting import ExcelReportGenerator, PDFReportGenerator
from wraith_crawler.services.auth import AuthService


def test_pdf_and_excel_reports(database, populated_assessment, tmp_path) -> None:
    pdf = tmp_path / "report.pdf"
    xlsx = tmp_path / "report.xlsx"
    with database.session() as session:
        PDFReportGenerator().generate(session, populated_assessment, pdf)
        ExcelReportGenerator().generate(session, populated_assessment, xlsx)
    assert pdf.stat().st_size > 5_000
    reader = PdfReader(str(pdf))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Attack Path Analysis" in text
    assert "Pentest Methodology" in text
    assert "Post-Exploitation Reasoning" in text
    assert "T1190" in text
    assert "Evidence Boundary" in text
    assert "SQL injection in search" in text
    workbook = load_workbook(xlsx, data_only=False)
    assert workbook.sheetnames == list(ExcelReportGenerator.SHEETS)
    assert workbook["Findings"]["D2"].value == "SQL injection in search"
    assert "T1190" in workbook["Findings"]["P2"].value
    assert workbook["Attack Paths"]["B2"].value == "SQL injection to sensitive data impact"
    assert "T1190" in workbook["Attack Path Findings"]["I2"].value
    assert workbook["Manual Review"]["D2"].value == "idor_bola"
    assert workbook["Findings"].freeze_panes == "A2"


async def login(client: httpx.AsyncClient) -> tuple[str, str]:
    response = await client.post("/api/v1/auth/login", json={"username": "administrator", "password": "StrongPassword123!"})
    assert response.status_code == 200
    return response.json()["session_token"], response.json()["csrf_token"]


async def test_api_auth_rbac_and_read_endpoints(database, config, admin, populated_assessment) -> None:
    transport = httpx.ASGITransport(app=create_app(config, database))
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        assert (await client.get("/health")).status_code == 200
        assert (await client.get("/api/v1/applications")).status_code == 401
        token, _ = await login(client)
        headers = {"Authorization": f"Bearer {token}"}
        me = await client.get("/api/v1/auth/me", headers=headers)
        assert me.status_code == 200
        assert "admin" in me.json()["roles"]
        response = await client.get(f"/api/v1/assessments/{populated_assessment}", headers=headers)
        assert response.status_code == 200
        assert response.json()["counts"]["findings"] == 1
        findings = await client.get(f"/api/v1/assessments/{populated_assessment}/findings", headers=headers)
        assert findings.json()[0]["priority_score"] == 96.0
        assert findings.json()[0]["mitre_attack"] == ["T1190"]
        paths = await client.get(f"/api/v1/assessments/{populated_assessment}/attack-paths", headers=headers)
        assert len(paths.json()) == 1
        assert paths.json()[0]["mitre_attack"] == ["T1190", "T1213.006"]
        for suffix in (
            "reconnaissance",
            "pentest-phases",
            "owasp-coverage",
            "post-exploitation",
        ):
            response = await client.get(
                f"/api/v1/assessments/{populated_assessment}/{suffix}", headers=headers
            )
            assert response.status_code == 200
        assert len((await client.get("/api/v1/owasp-coverage", headers=headers)).json()) == 10
        mitre = await client.get("/api/v1/mitre-attack", headers=headers)
        assert mitre.status_code == 200
        assert any(item["technique_id"] == "T1190" for item in mitre.json())


async def test_cookie_mutation_requires_csrf(database, config, admin) -> None:
    transport = httpx.ASGITransport(app=create_app(config, database))
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        _, csrf = await login(client)
        response = await client.post(
            "/api/v1/users",
            json={"username": "seconduser", "password": "StrongPassword123!", "roles": ["viewer"]},
        )
        assert response.status_code == 403
        response = await client.post(
            "/api/v1/users",
            headers={"X-CSRF-Token": csrf},
            json={"username": "seconduser", "password": "StrongPassword123!", "roles": ["viewer"]},
        )
        assert response.status_code == 201


async def test_viewer_cannot_create_users(database, config) -> None:
    with database.session() as session:
        AuthService().create_user(session, "readonly", "StrongPassword123!", [__import__("wraith_crawler.enums", fromlist=["RoleName"]).RoleName.VIEWER])
    transport = httpx.ASGITransport(app=create_app(config, database))
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        response = await client.post("/api/v1/auth/login", json={"username": "readonly", "password": "StrongPassword123!"})
        token = response.json()["session_token"]
        denied = await client.post(
            "/api/v1/users",
            headers={"Authorization": f"Bearer {token}"},
            json={"username": "nopeuser", "password": "StrongPassword123!", "roles": ["viewer"]},
        )
        assert denied.status_code == 403


async def test_api_report_output_cannot_escape_configured_root(database, config, admin, populated_assessment) -> None:
    transport = httpx.ASGITransport(app=create_app(config, database))
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        token, _ = await login(client)
        response = await client.post(
            f"/api/v1/assessments/{populated_assessment}/reports",
            headers={"Authorization": f"Bearer {token}"},
            json={"report_type": "pdf", "output_directory": "/private/tmp/outside-wraith-output"},
        )
        assert response.status_code == 422
